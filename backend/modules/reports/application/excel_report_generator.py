"""Excel (.xlsx) generation for Reports exports, using openpyxl. One sheet
for the KPI summary, one sheet per breakdown table -- so the export stays
usable as raw data (pivot-table friendly) rather than a printed layout."""
import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from modules.reports.application.export_adapter import ExportSections

_HEADER_FILL = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_TITLE_FONT = Font(size=14, bold=True, color="1E293B")


def _autosize(ws) -> None:
    for col_cells in ws.columns:
        length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max(length + 2, 10), 40)


def generate_report_excel(*, sections: ExportSections, company_name: str) -> bytes:
    wb = Workbook()

    summary = wb.active
    summary.title = "Summary"
    summary["A1"] = f"{company_name} — {sections.title}"
    summary["A1"].font = _TITLE_FONT
    summary["A2"] = f"Period: {sections.date_from} to {sections.date_to}"

    summary.append([])
    summary.append(["Metric", "Value"])
    for cell in summary[summary.max_row]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
    for label, value in sections.kpis:
        summary.append([label, value])
    _autosize(summary)

    for table in sections.tables:
        sheet_name = table.title[:31]  # Excel's hard sheet-name length limit.
        ws = wb.create_sheet(title=sheet_name)
        ws.append(table.headers)
        for cell in ws[1]:
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = Alignment(horizontal="left")
        for row in table.rows:
            ws.append(row)
        _autosize(ws)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
